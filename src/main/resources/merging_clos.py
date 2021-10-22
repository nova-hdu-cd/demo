# coding=utf-8
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment
file=os.path.dirname(os.path.abspath(__file__))+os.sep+"ActivityRoleExport.xlsx"
df = pd.read_excel(file)
wb = load_workbook(file)
ws=wb.active

# print (df)
grous = df.groupby('角色').groups
# nums=df.groupby('城市号').groups
aligncenter = Alignment(horizontal='center', vertical='center')

for grou in grous:
    indexs=grous[grou]+2
    indexs = indexs.sort_values()
    # print (indexs)
    # print (indexs[0],indexs[-1])
    ws.merge_cells(start_row=indexs[0],end_row=indexs[-1],start_column=1,end_column=1)
    ws.merge_cells(start_row=indexs[0],end_row=indexs[-1],start_column=2,end_column=2)
    for i in range(1, ws.max_row + 1):  # 实现居中
        ws.cell(row=i, column=1).alignment = aligncenter
        ws.cell(row=i, column=2).alignment = aligncenter




wb.save(os.path.dirname(os.path.abspath(__file__))+os.sep+'hebin.xlsx')